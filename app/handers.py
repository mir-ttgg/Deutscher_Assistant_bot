import io
import logging
from os import getenv
from pathlib import Path

import openpyxl
from aiogram import F, Router
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, Message

from app.ai import get_ai_response, update_knowledge_base
from app.keyboards import (get_article_keyboards, get_back_to_admin_keyboard,
                           get_start_admin_keyboard,
                           get_user_management_keyboard)
from datebase.config import AsyncSessionLocal, generate_users_report
from datebase.crud import (add_user, delete_user, get_all_queries,
                           get_user_by_id, increment_user_count, is_blocked,
                           save_query)


class AddUser(StatesGroup):
    waiting_for_user_id = State()


class DeleteUser(StatesGroup):
    waiting_for_user_id = State()


class UpdateKnowledge(StatesGroup):
    waiting_for_file = State()


router = Router()
MANAGER_ID = [int(x) for x in getenv('MANAGER_ID', '').split(',') if x]

TOPICS_MESSAGE = """К сожалению, я не могу проконсультировать вас по этой теме. Я отвечаю только на вопросы по темам:

1️⃣ Прописка
2️⃣ Банковская карта и счет
3️⃣ Налоговый номер
4️⃣ SIM-карта и мобильная связь
5️⃣ Общественный транспорт
6️⃣ Медицинская система и запись к врачам
7️⃣ Аптеки и лекарства
8️⃣ Почта и официальная корреспонденция
9️⃣ Коммуникация с Ausländerbehörde
1️⃣0️⃣ Мусор и система сортировки
1️⃣1️⃣ Тишина, соседи и соблюдение правил проживания
1️⃣2️⃣ Супермаркеты и торговые сети"""


@router.message(CommandStart())
async def start(message: Message):

    if message.from_user.id in MANAGER_ID:
        await message.answer('Выберите действие', reply_markup=get_start_admin_keyboard())

    else:
        try:
            user = await get_user_by_id(message.from_user.id)
            if user is None:
                await message.answer('❌ У вас нет доступа. Пожалуйста, обратитесь к своему куратору для получения доступа к сервису.')
                return
            else:
                username = message.from_user.username or 'unknown'
                await add_user(message.from_user.id, username)
                if await is_blocked(message.from_user.id):
                    await message.answer('К сожалению, вы исчерпали лимит вопросов на эту неделю. Счетчик обновляется по понедельникам.')
                    return
                else:
                    await message.answer('''👋 Привет! Я твой ИИ-помощник по адаптации в Германии. Я могу проконсультировать тебя по темам: 
1️⃣ Прописка
2️⃣ Банковская карта и счет
3️⃣ Налоговый номер
4️⃣ SIM-карта и мобильная связь
5️⃣ Общественный транспорт
6️⃣ Медицинская система и запись к врачам
7️⃣ Аптеки и лекарства
8️⃣ Почта и официальная корреспонденция
9️⃣ Коммуникация с Ausländerbehörde
1️⃣0️⃣ Мусор и система сортировки
1️⃣1️⃣ Тишина, соседи и соблюдение правил проживания
1️⃣2️⃣ Супермаркеты и торговые сети
''')
        except Exception as e:
            logging.info(f'Ошибка получения пользователя: {e}')


@router.message(F.from_user.id.in_(MANAGER_ID), F.text == 'Управление пользователями')
async def manage_users(message: Message):
    await message.answer('Выберите действие:', reply_markup=get_user_management_keyboard())


def is_valid_user_id_format(value: str) -> bool:
    '''
    Проверяет, что строка является целым положительным числом.
    '''
    try:
        user_id = int(value)
        return user_id > 0
    except (ValueError, TypeError):
        return False


@router.callback_query(F.data == 'back_admin')
async def back_to_admin(callback_query: CallbackQuery):
    await callback_query.message.delete()
    await callback_query.message.answer(
        'Вы в главном меню:',
        reply_markup=get_start_admin_keyboard()
    )
    await callback_query.answer()


async def back_to_admin_msg(message):
    await message.answer(
        'Вы в главном меню:',
        reply_markup=get_start_admin_keyboard()
    )


@router.callback_query(F.data == 'add_user')
async def add_user_callback(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_text(
        text='Введите Telegram ID пользователя',
        reply_markup=None
    )
    # await callback_query.message.answer('Введите Telegram ID пользователя')
    await state.set_state(AddUser.waiting_for_user_id)


@router.message(AddUser.waiting_for_user_id)
async def add_user_receive(message: Message, state: FSMContext):
    try:
        user_id = int(message.text)
        if not is_valid_user_id_format(str(user_id)):
            await message.answer('Неверный формат Telegram ID. Пожалуйста, введите положительное целое число.')
            await back_to_admin_msg(message)
            await state.clear()
            return
        await add_user(user_id, 'unknown')
        logging.info(
            f'Пользователь {user_id} добавлен менеджером {message.from_user.id}')
        await message.answer('Пользователь добавлен')
        await back_to_admin_msg(message)
        await state.clear()
        return
    except Exception as e:
        logging.info(f'Ошибка добавления пользователя: {e}')
        await message.answer('Неверный формат Telegram ID. Пожалуйста, введите положительное целое число.')
        await back_to_admin_msg(message)
        await state.clear()
        return


@router.callback_query(F.data == 'list_users')
async def list_users_callback(callback_query: CallbackQuery, session: AsyncSessionLocal):
    try:
        file_stream = await generate_users_report(session)

        # Формируем имя файла с датой
        filename = 'users_list.xlsx'
        loading_msg = await callback_query.message.edit_text(
            text='Формирую отчёт, пожалуйста, подождите...',
            reply_markup=None
        )
        # Создаем объект файла для отправки
        document = BufferedInputFile(
            file=file_stream.read(),
            filename=filename
        )

        # Отправляем файл
        await callback_query.message.answer_document(
            document=document,
            caption="📋 Список пользователей"
        )
        await loading_msg.delete()
        await callback_query.answer('Отчёт сформирован')
    except Exception as e:
        logging.error(f'Ошибка при генерации отчёта: {e}')
        await callback_query.message.answer('Не удалось сформировать отчёт')
        await callback_query.answer()


@router.callback_query(F.data == 'delete_user')
async def delete_user_callback(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_text(
        text='Введите Telegram ID пользователя для удаления',
        reply_markup=None
    )
    await state.set_state(DeleteUser.waiting_for_user_id)


@router.message(DeleteUser.waiting_for_user_id)
async def delete_user_receive(message: Message, state: FSMContext):
    try:
        user_id = int(message.text)
        if not is_valid_user_id_format(str(user_id)):
            await message.answer('Неверный формат Telegram ID. Пожалуйста, введите положительное целое число.')
            await back_to_admin_msg(message)
            await state.clear()
            return
        if await delete_user(user_id):
            logging.info(
                f'Пользователь {user_id} удален менеджером {message.from_user.id}')
            await message.answer('Пользователь удален')
            await back_to_admin_msg(message)
            await state.clear()
            return
        else:
            await message.answer('Пользователь не найден')
            await back_to_admin_msg(message)
            await state.clear()
            return
    except Exception as e:
        logging.info(f'Ошибка удаления пользователя: {e}')
        await message.answer('Неверный формат Telegram ID. Пожалуйста, введите положительное целое число.')
        await back_to_admin_msg(message)
        await state.clear()
        return


@router.message(F.from_user.id.in_(MANAGER_ID), F.text == 'Выгрузка отчёта')
async def export_report_callback(message: Message):
    """Выгрузка отчета со всеми вопросами и ответами"""
    try:
        load_mgs = await message.answer('Формирую отчет, пожалуйста подождите...')

        queries = await get_all_queries()

        if not queries:
            await message.answer('❌ Нет данных для отчета')
            await load_mgs.delete()
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Вопросы и ответы"

        ws['A1'] = 'User ID'
        ws['B1'] = 'Дата (МСК)'
        ws['C1'] = 'Вопрос пользователя'
        ws['D1'] = 'Ответ ИИ'

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for idx, query in enumerate(queries, start=2):
            ws[f'A{idx}'] = query.user_id
            if query.created_at.tzinfo is None:
                ws[f'B{idx}'] = query.created_at
            else:
                from zoneinfo import ZoneInfo
                msk_time = query.created_at.astimezone(
                    ZoneInfo("Europe/Moscow"))
                ws[f'B{idx}'] = msk_time.replace(tzinfo=None)

            ws[f'C{idx}'] = query.question
            ws[f'D{idx}'] = query.answer

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 70

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from datetime import datetime
        filename = f'queries_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        document = BufferedInputFile(
            file=buffer.read(),
            filename=filename
        )

        await message.answer_document(
            document=document,
            caption=f"Отчет по вопросам и ответам\nВсего записей: {len(queries)}"
        )
        await load_mgs.delete()
    except Exception as e:
        await load_mgs.delete()
        logging.error(f'Ошибка при генерации отчета: {e}')
        await message.answer('❌ Не удалось сформировать отчет')


@router.message(F.from_user.id.in_(MANAGER_ID), F.text == 'База знаний')
async def knowledge_base_callback(message: Message):
    """Показываем текущую базу знаний"""
    try:
        kb_path = Path('knowledge/document1.docx')

        if kb_path.exists():
            with open(kb_path, 'rb') as f:
                document = BufferedInputFile(
                    file=f.read(),
                    filename='current_knowledge_base.docx'
                )

            await message.answer_document(
                document=document,
                caption="📚 Вот текущая база знаний",
                reply_markup=get_article_keyboards()
            )
        else:
            await message.answer(
                "❌ База знаний не найдена",
                reply_markup=get_article_keyboards()
            )
    except Exception as e:
        logging.error(f'Ошибка при отправке БЗ: {e}')
        await message.answer("❌ Ошибка при загрузке базы знаний")


@router.callback_query(F.data == 'edit_article')
async def edit_article_callback(callback_query: CallbackQuery, state: FSMContext):
    """Начало процесса обновления БЗ"""

    await callback_query.message.delete()

    # Отправляем новое
    await callback_query.message.answer(
        text='📤 Отправьте новую БЗ в формате .docx',
        reply_markup=get_back_to_admin_keyboard()
    )
    await state.set_state(UpdateKnowledge.waiting_for_file)
    await callback_query.answer()


@router.message(UpdateKnowledge.waiting_for_file, F.document)
async def receive_new_knowledge_base(message: Message, state: FSMContext):
    """Обработка загрузки новой БЗ"""
    try:
        # Проверяем формат файла
        if not message.document.file_name.endswith('.docx'):
            await message.answer('❌ К сожалению не получилось загрузить новую БЗ, проверьте формат файла!')
            await back_to_admin_msg(message)
            await state.clear()
            return

        msg = await message.answer('⏳ Обрабатываю файл...')

        # Скачиваем файл
        file = await message.bot.get_file(message.document.file_id)
        file_path = Path('knowledge/document1.docx')
        file_path.parent.mkdir(exist_ok=True)

        await message.bot.download_file(file.file_path, file_path)

        # Обновляем базу знаний в OpenAI
        await update_knowledge_base(str(file_path))

        logging.info(f'БЗ обновлена менеджером {message.from_user.id}')
        await message.answer('✅ Новая БЗ загружена успешно!')
        await msg.delete()
        await back_to_admin_msg(message)
        await state.clear()

    except Exception as e:
        logging.error(f'Ошибка обновления БЗ: {e}')
        await msg.delete()
        await message.answer('❌ К сожалению не получилось загрузить новую БЗ, проверьте формат файла!')
        await back_to_admin_msg(message)
        await state.clear()


@router.message(F.content_type.in_(['voice', 'sticker', 'photo', 'video', 'document', 'audio', 'video_note']))
async def handle_non_text(message: Message):
    """Обработка не текстовых сообщений"""
    user = await get_user_by_id(message.from_user.id)

    if user is None and message.from_user.id not in MANAGER_ID:
        return

    await message.answer(
        "Прости, я умею читать только текстовые сообщения, "
        "я не умею слушать голосовые сообщения или читать файлы отправленные тобой. "
        "Напиши свой вопрос текстом я обязательно помогу тебе!"
    )


@router.message(F.text)
async def handle_question(message: Message):
    """Обработка текстовых вопросов пользователей"""

    if message.from_user.id in MANAGER_ID:
        return

    # Проверяем доступ
    user = await get_user_by_id(message.from_user.id)
    if user is None:
        await message.answer('❌ У вас нет доступа. Пожалуйста, обратитесь к своему куратору для получения доступа к сервису.')
        return

    if await is_blocked(message.from_user.id):
        await message.answer('К сожалению, вы исчерпали лимит вопросов на эту неделю. Счетчик обновляется по понедельникам.')
        return

    await message.bot.send_chat_action(message.chat.id, 'typing')

    try:
        response = await get_ai_response(message.text)

        await increment_user_count(message.from_user.id)

        await save_query(message.from_user.id, message.text, response)
        updated_user = await get_user_by_id(message.from_user.id)
        remaining = 100 - updated_user.count

        await message.answer(response)

        logging.info(
            f'Пользователь {message.from_user.id} задал вопрос. Осталось запросов: {remaining}')

    except Exception as e:
        logging.error(f'Ошибка при обработке вопроса: {e}')
        await message.answer('😔 Произошла ошибка при обработке вашего вопроса. Попробуйте еще раз.')
